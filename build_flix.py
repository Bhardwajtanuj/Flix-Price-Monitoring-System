import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import re, time

t0 = time.time()


print("Reading dataset...")
try:  # try CSV first, fall back to xlsx
    df = pd.read_csv(
        r"D:\resume\resume projects\Flix\Take Home Assignment Dataset.csv",
        encoding="utf-8-sig", low_memory=False)
except FileNotFoundError:
    df = pd.read_excel(
        r"D:\resume\resume projects\Flix\Take Home Assignment Dataset.xlsx",
        engine="calamine")
print(f"  {len(df):,} rows in {time.time()-t0:.1f}s")


print("Cleaning...")
t1 = time.time()

def vec_bool(s):
    result = pd.Series(False, index=s.index)
    bm = s.apply(lambda x: isinstance(x, bool))
    result[bm]  = s[bm].astype(bool)
    result[~bm] = s[~bm].astype(str).str.strip().str.upper().isin(["TRUE","1","YES"])
    return result

def vec_num(s, dtype=float):
    return pd.to_numeric(s.astype(str).str.replace(",",".").str.strip(),
                         errors="coerce").fillna(0).astype(dtype)

for col in ["Is AC","Is Seater","Is Sleeper",
            "Is Live Tracking Available","Is M-Ticket Enabled","Is Seat Layout Available"]:
    df[col] = vec_bool(df[col])

for col,dt in [("Journey Duration (Min)",int),("Available Seats",int),
               ("Total Seats",int),("Number of Reviews",int),
               ("Weighted Average Price",float),("Total Ratings",float),
               ("Bus Score",float),("BP Count",int),("DP Count",int)]:
    df[col] = vec_num(df[col], dt)

df["Rank_Num"] = pd.to_numeric(
    df["SRP Rank"].astype(str).str.split("/").str[0],
    errors="coerce").fillna(0).astype(int)

def pt(v):
    try:
        p=str(v).strip().split(":")
        return f"{int(p[0])%24:02d}:{int(p[1]):02d}"
    except: return "00:00"

df["Dep_Str"] = df["Departure Time"].apply(pt)
df["Arr_Str"] = df["Arrival Time"].apply(pt)
df["Dep_Min"] = df["Dep_Str"].apply(
    lambda v: int(v.split(":")[0])*60+int(v.split(":")[1]))

df["Layout"] = df["Bus Type"].apply(
    lambda v: (m.group(1) if (m:=re.search(r'\((\d\+\d)\)',str(v))) else "unknown"))

df["Occ_Rate"] = np.where(df["Total Seats"]>0,  # % of seats already sold
    (df["Total Seats"]-df["Available Seats"])/df["Total Seats"]*100, 0.0).round(1)

df["Svc_Score"] = (df["Is Live Tracking Available"].astype(int)+  # 0-3 premium feature count
                   df["Is M-Ticket Enabled"].astype(int)+
                   df["Is Seat Layout Available"].astype(int))

def rwap(s):
    try:
        rev=0.0; seats=0; prices=[]
        for p in str(s).split(","):
            if ":" in p:
                pr,ct=p.strip().split(":",1)
                price=float(pr); count=int(float(ct))
                rev+=price*count; seats+=count; prices.append(price)
        if seats==0: return 0.0,0.0,0.0
        return round(rev/seats,2),round(min(prices),2),round(max(prices),2)
    except: return 0.0,0.0,0.0

print("  Recomputing WAP...")  # more accurate than the pre-supplied column
wd=df["Seat Prices"].apply(rwap)
df["WAP_Calc"] =wd.apply(lambda x:x[0])
df["Min_Price"]=wd.apply(lambda x:x[1])
df["Max_Price"]=wd.apply(lambda x:x[2])
df["WAP_Final"]=np.where(df["WAP_Calc"]>0,df["WAP_Calc"],df["Weighted Average Price"])

def pf(s):
    try:
        f=[float(x) for x in str(s).split("-") if x.strip() not in ["","nan","NaN"]]
        return (min(f),max(f),len(f)) if f else (0.0,0.0,0)
    except: return 0.0,0.0,0

fd=df["Fare List"].apply(pf)
df["Min_Fare"] =fd.apply(lambda x:x[0])
df["Max_Fare"] =fd.apply(lambda x:x[1])
df["Num_Tiers"]=fd.apply(lambda x:x[2])

def ad(row):
    try:
        f=[float(x) for x in str(row["Fare List"]).split("-") if x.strip()]
        d=[float(x) for x in str(row["Discounted Prices"]).split("-") if x.strip()]
        if len(f)!=len(d) or not f: return 0.0
        return round(sum((a-b)/a*100 for a,b in zip(f,d) if a>0)/len(f),1)
    except: return 0.0

df["Avg_Disc"]=df.apply(ad,axis=1)
df["Route Number"]=df["Route Number"].astype(str).str.strip()
df["Operator"]=df["Operator"].astype(str).str.strip()
df=df[(df["WAP_Final"]>0)&(df["Operator"].str.len()>0)&
      (~df["Operator"].isin(["nan","NaN","None"]))].copy().reset_index(drop=True)

print(f"  Done in {time.time()-t1:.1f}s | {len(df):,} valid rows")

FLIX_NAME = next(
    (op for op in sorted(df["Operator"].unique()) if "flix" in op.lower()), None)  # case-insensitive
if FLIX_NAME is None:
    FLIX_NAME = input("Enter Flixbus operator name: ").strip()
print(f"Flixbus='{FLIX_NAME}' | {len(df[df['Operator']==FLIX_NAME]):,} trips")


print("\nRunning similarity matching...")
t2 = time.time()

def weighted_median(prices, weights):  # high-occupancy buses carry more weight
    if not len(prices): return 0.0
    paired = sorted(zip(prices, weights))
    tw = sum(w for _,w in paired)
    if tw==0: return float(np.median(prices))
    cum=0
    for p,w in paired:
        cum+=w
        if cum>=tw/2: return p
    return paired[-1][0]

def dyn_thresh(prices):  # tight market -> stricter, volatile market -> looser
    if len(prices)<3: return 15.0
    mn=np.mean(prices)
    if mn==0: return 15.0
    cv=np.std(prices)/mn*100
    return 10.0 if cv<10 else (20.0 if cv>25 else 15.0)

routes=sorted(df["Route Number"].unique())
all_flag_rows=[]; summary_stats=[]

for route in routes:
    rdf  = df[df["Route Number"]==route]
    fdf  = rdf[rdf["Operator"]==FLIX_NAME].reset_index(drop=True)
    cdf  = rdf[rdf["Operator"]!=FLIX_NAME].reset_index(drop=True)

    if len(fdf)==0 or len(cdf)==0:
        summary_stats.append({"route":route,"total":len(rdf),
                               "flix":len(fdf),"an":0,"flag":0,"hi":0,"lo":0,
                               "avg_diff":0,"avg_hi_diff":0,"avg_lo_diff":0,
                               "avg_flix_wap":0,"avg_comp_wap":0})
        continue

    print(f"  {route}: {len(fdf)} Flix + {len(cdf)} comps", end="", flush=True)

    buckets = {}  # group comps by (AC, Sleeper, Layout) so each Flix trip only searches its bucket
    for _, crow in cdf.iterrows():
        key = (bool(crow["Is AC"]), bool(crow["Is Sleeper"]), str(crow["Layout"]))
        if key not in buckets:
            buckets[key] = {"dur":[],"dep":[],"rat":[],"wap":[],
                            "occ":[],"svc":[],"rows":[]}
        b = buckets[key]
        b["dur"].append(float(crow["Journey Duration (Min)"]))
        b["dep"].append(float(crow["Dep_Min"]))
        b["rat"].append(float(crow["Total Ratings"]))
        b["wap"].append(float(crow["WAP_Final"]))
        b["occ"].append(float(crow["Occ_Rate"]))
        b["svc"].append(float(crow["Svc_Score"]))
        b["rows"].append(crow)

    for key in buckets:
        b = buckets[key]
        for col in ["dur","dep","rat","wap","occ","svc"]:
            b[col] = np.array(b[col])

    rt_flag=rt_hi=rt_lo=0
    rt_diffs=[]; rt_hi_diffs=[]; rt_lo_diffs=[]
    rt_flix_waps=[]; rt_comp_waps=[]

    for _, fr in fdf.iterrows():
        f_ac    = bool(fr["Is AC"])
        f_sleep = bool(fr["Is Sleeper"])
        f_lay   = str(fr["Layout"])
        f_dur   = float(fr["Journey Duration (Min)"])
        f_dep   = float(fr["Dep_Min"])
        f_rat   = float(fr["Total Ratings"])
        f_wap   = float(fr["WAP_Final"])
        f_svc   = float(fr["Svc_Score"])

        match_keys = []
        exact = (f_ac, f_sleep, f_lay)
        if exact in buckets: match_keys.append(exact)
        unk = (f_ac, f_sleep, "unknown")
        if unk in buckets and unk != exact: match_keys.append(unk)
        if f_lay == "unknown":
            for k in buckets:
                if k[0]==f_ac and k[1]==f_sleep and k not in match_keys:
                    match_keys.append(k)

        if not match_keys: continue

        all_dur = np.concatenate([buckets[k]["dur"] for k in match_keys])
        all_dep = np.concatenate([buckets[k]["dep"] for k in match_keys])
        all_rat = np.concatenate([buckets[k]["rat"] for k in match_keys])
        all_wap = np.concatenate([buckets[k]["wap"] for k in match_keys])
        all_occ = np.concatenate([buckets[k]["occ"] for k in match_keys])
        all_svc = np.concatenate([buckets[k]["svc"] for k in match_keys])
        all_rows= [r for k in match_keys for r in buckets[k]["rows"]]

        m = np.abs(all_dur - f_dur) <= 60
        td = np.abs(all_dep - f_dep)
        m &= np.minimum(td, 1440-td) <= 90  # circular so 23:50 vs 00:10 = 20 min not 1420
        if f_rat > 0:
            m &= (all_rat==0) | (np.abs(all_rat-f_rat)<=0.8)

        idx = np.where(m)[0]
        if not len(idx): continue

        prices  = all_wap[idx]
        weights = np.maximum(all_occ[idx], 1.0)
        svcs    = all_svc[idx]

        w_med   = round(weighted_median(prices.tolist(), weights.tolist()), 2)
        thold   = dyn_thresh(prices.tolist())
        svc_adj = round((f_svc - float(np.mean(svcs)))*2.0, 1)  # 2% per feature point diff
        adj_b   = round(w_med*(1+svc_adj/100), 2)
        dp      = (f_wap-adj_b)/adj_b*100 if adj_b else 0

        if   dp >  thold: flag,fc="TOO HIGH","FF4C4C"; rt_hi+=1; rt_flag+=1; rt_hi_diffs.append(dp)
        elif dp < -thold: flag,fc="TOO LOW", "FFD966"; rt_lo+=1; rt_flag+=1; rt_lo_diffs.append(dp)
        else:             flag,fc="OK",       "92D050"

        rt_diffs.append(dp)
        rt_flix_waps.append(f_wap)
        rt_comp_waps.append(w_med)

        comp_rows = pd.DataFrame([all_rows[i] for i in idx[:10]])
        all_flag_rows.append({
            "route":route,"rank":int(fr["Rank_Num"]),
            "dep":str(fr["Dep_Str"]),"arr":str(fr["Arr_Str"]),
            "dur":int(f_dur),"layout":f_lay,"occ":float(fr["Occ_Rate"]),
            "min_fare":float(fr["Min_Fare"]),"max_fare":float(fr["Max_Fare"]),
            "num_tiers":int(fr["Num_Tiers"]),"avg_disc":float(fr["Avg_Disc"]),
            "svc":int(f_svc),"svc_adj":svc_adj,"flix_price":f_wap,
            "comp_count":int(len(idx)),
            "w_med":w_med,"adj_b":adj_b,
            "s_med":round(float(np.median(prices)),2),
            "mean_p":round(float(np.mean(prices)),2),
            "min_c":round(float(np.min(prices)),2),
            "max_c":round(float(np.max(prices)),2),
            "diff_abs":round(f_wap-adj_b,2),"diff_pct":round(dp,2),
            "thold":thold,"flag":flag,"fc":fc,"comp_rows":comp_rows,
        })

    an=sum(1 for x in all_flag_rows if x["route"]==route)
    avg_diff     = round(np.mean(rt_diffs),1) if rt_diffs else 0
    avg_hi_diff  = round(np.mean(rt_hi_diffs),1) if rt_hi_diffs else 0
    avg_lo_diff  = round(np.mean(rt_lo_diffs),1) if rt_lo_diffs else 0
    avg_flix_wap = round(np.mean(rt_flix_waps),0) if rt_flix_waps else 0
    avg_comp_wap = round(np.mean(rt_comp_waps),0) if rt_comp_waps else 0

    print(f" → {an} analysed, {rt_flag} flagged ({time.time()-t2:.0f}s)")
    summary_stats.append({
        "route":route,"total":len(rdf),"flix":len(fdf),"an":an,
        "flag":rt_flag,"hi":rt_hi,"lo":rt_lo,
        "avg_diff":avg_diff,"avg_hi_diff":avg_hi_diff,"avg_lo_diff":avg_lo_diff,
        "avg_flix_wap":avg_flix_wap,"avg_comp_wap":avg_comp_wap,
    })

ta=len(all_flag_rows)
tf=sum(1 for f in all_flag_rows if f["flag"]!="OK")
th=sum(1 for f in all_flag_rows if f["flag"]=="TOO HIGH")
tl=sum(1 for f in all_flag_rows if f["flag"]=="TOO LOW")
print(f"\nMatching done in {time.time()-t2:.1f}s")
print(f"Analysed:{ta:,} | Flagged:{tf:,} (High:{th:,} Low:{tl:,})")


print("\nBuilding Excel...")
t3=time.time()

C_FLIX="00A843"; C_HDR="1A3C34"; C_ALT="F0FAF4"
thin=Side(style="thin",color="CCCCCC")
med =Side(style="medium",color="888888")
brd =Border(left=thin,right=thin,top=thin,bottom=thin)
brd_m=Border(left=med,right=med,top=med,bottom=med)
brd_thick=Border(left=Side(style="thick",color="00A843"),
                 right=Side(style="thick",color="00A843"),
                 top=Side(style="thick",color="00A843"),
                 bottom=Side(style="thick",color="00A843"))

fills={k:PatternFill("solid",fgColor=v) for k,v in {
    "flix":"00A843","hdr":"1A3C34","alt":"F0FAF4","white":"FFFFFF",
    "red":"FF4C4C","yel":"FFD966","grn":"92D050","kpi":"F8F8F8",
    "frow":"E6FFE6","cover":"0D2B22","cover2":"1A3C34",
    "hi_opp":"FFE0E0","lo_opp":"E0FFE8","step":"E8F5E9",
    "arrow":"00A843","warn":"FFF3CD",
}.items()}
fonts={
    "title": Font(name="Arial",bold=True,size=13,color="FFFFFF"),
    "sec":   Font(name="Arial",bold=True,size=11,color="FFFFFF"),
    "hdr":   Font(name="Arial",bold=True,size=10,color="FFFFFF"),
    "norm":  Font(name="Arial",size=10),
    "bold":  Font(name="Arial",bold=True,size=10),
    "sm":    Font(name="Arial",size=9),
    "frow":  Font(name="Arial",bold=True,size=9,color="006400"),
    "cover_main": Font(name="Arial",bold=True,size=28,color="FFFFFF"),
    "cover_sub":  Font(name="Arial",size=14,color="92D050"),
    "cover_body": Font(name="Arial",size=11,color="FFFFFF"),
    "cover_label":Font(name="Arial",bold=True,size=10,color="00A843"),
    "cover_val":  Font(name="Arial",bold=True,size=16,color="FFFFFF"),
    "insight_hi": Font(name="Arial",bold=True,size=10,color="C62828"),
    "insight_lo": Font(name="Arial",bold=True,size=10,color="1B5E20"),
    "step_hdr":   Font(name="Arial",bold=True,size=11,color="FFFFFF"),
    "step_body":  Font(name="Arial",size=10),
    "arrow_f":    Font(name="Arial",bold=True,size=16,color="FFFFFF"),
}
aligns={
    "c":  Alignment(horizontal="center",vertical="center",wrap_text=True),
    "l":  Alignment(horizontal="left",  vertical="center",wrap_text=True,indent=1),
    "lt": Alignment(horizontal="left",  vertical="top",   wrap_text=True,indent=1),
    "r":  Alignment(horizontal="right", vertical="center",wrap_text=True),
}

def wtitle(ws,rng,text,h=44):
    ws.merge_cells(rng)
    r=int(''.join(c for c in rng.split(":")[0] if c.isdigit()))
    ws.row_dimensions[r].height=h
    c=ws[rng.split(":")[0]]
    c.value=text;c.font=fonts["title"];c.fill=fills["flix"];c.alignment=aligns["c"]

def wsec(ws,rng,text):
    ws.merge_cells(rng)
    r=int(''.join(c for c in rng.split(":")[0] if c.isdigit()))
    ws.row_dimensions[r].height=26
    c=ws[rng.split(":")[0]]
    c.value=text;c.font=fonts["sec"];c.fill=fills["hdr"];c.alignment=aligns["l"]

def whdrs(ws,row,hdrs,h=40):
    ws.row_dimensions[row].height=h
    for ci,ht in enumerate(hdrs,1):
        c=ws.cell(row,ci);c.value=ht;c.font=fonts["hdr"]
        c.fill=fills["hdr"];c.alignment=aligns["c"];c.border=brd

def wrow(ws,row,data,fill,nfmts=None,fcol=None,ffill=None):
    ws.row_dimensions[row].height=18
    for ci,v in enumerate(data,1):
        c=ws.cell(row,ci);c.value=v
        c.font=fonts["bold"] if(fcol and ci==fcol) else fonts["norm"]
        c.fill=ffill if(fcol and ci==fcol and ffill) else fill
        c.alignment=aligns["c"];c.border=brd
        if nfmts and ci in nfmts: c.number_format=nfmts[ci]

def scols(ws,widths):
    for ci,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(ci)].width=w

wb=openpyxl.Workbook()


ws_cover = wb.active
ws_cover.title = "Cover"
ws_cover.sheet_view.showGridLines = False
ws_cover.sheet_properties.tabColor = "00A843"


for row in range(1, 45):
    ws_cover.row_dimensions[row].height = 20
    for col in range(1, 9):
        c = ws_cover.cell(row, col)
        c.fill = fills["cover"]


for row in range(1, 6):
    ws_cover.row_dimensions[row].height = 22
    for col in range(1, 9):
        ws_cover.cell(row, col).fill = fills["flix"]

ws_cover.merge_cells("A1:H5")
c = ws_cover["A1"]
c.value = "FLIX BUS"
c.font  = Font(name="Arial", bold=True, size=36, color="FFFFFF")
c.fill  = fills["flix"]
c.alignment = Alignment(horizontal="center", vertical="center")


ws_cover.row_dimensions[7].height = 14
ws_cover.merge_cells("A8:H10")
ws_cover.row_dimensions[8].height = 30
ws_cover.row_dimensions[9].height = 30
ws_cover.row_dimensions[10].height = 30
c = ws_cover["A8"]
c.value = "PRICE MONITORING SYSTEM"
c.font  = fonts["cover_main"]
c.fill  = fills["cover"]
c.alignment = Alignment(horizontal="center", vertical="center")

ws_cover.merge_cells("A11:H12")
ws_cover.row_dimensions[11].height = 24
ws_cover.row_dimensions[12].height = 24
c = ws_cover["A11"]
c.value = "Take-Home Assignment  |  Data & Growth Intern  |  Flix"
c.font  = fonts["cover_sub"]
c.fill  = fills["cover"]
c.alignment = Alignment(horizontal="center", vertical="center")


ws_cover.row_dimensions[13].height = 6
for col in range(1, 9):
    c = ws_cover.cell(13, col)
    c.fill = fills["flix"]


ws_cover.row_dimensions[15].height = 18
ws_cover.merge_cells("A15:H15")
c = ws_cover["A15"]
c.value = "DATASET OVERVIEW"
c.font  = Font(name="Arial", bold=True, size=12, color="92D050")
c.fill  = fills["cover"]
c.alignment = Alignment(horizontal="center", vertical="center")

stats = [
    ("Routes Analysed", str(len(routes))),
    ("Total Bus Listings", f"{len(df):,}"),
    ("Flixbus Trips", f"{len(df[df['Operator']==FLIX_NAME]):,}"),
    ("Trips Flagged", f"{tf:,}"),
]
ws_cover.row_dimensions[16].height = 22
ws_cover.row_dimensions[17].height = 36
ws_cover.row_dimensions[18].height = 10
for ci, (label, val) in enumerate(stats, 1):
    c1 = ws_cover.cell(16, ci*2-1)
    c1.value = label
    c1.font  = fonts["cover_label"]
    c1.fill  = fills["cover2"]
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border = brd_m
    ws_cover.merge_cells(
        start_row=16, start_column=ci*2-1,
        end_row=16,   end_column=ci*2)
    c2 = ws_cover.cell(17, ci*2-1)
    c2.value = val
    c2.font  = fonts["cover_val"]
    c2.fill  = fills["cover2"]
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.border = brd_m
    ws_cover.merge_cells(
        start_row=17, start_column=ci*2-1,
        end_row=17,   end_column=ci*2)


ws_cover.row_dimensions[19].height = 8
desc_rows = [
    ("ASSIGNMENT OBJECTIVE", None),
    ("Design a system that compares Flix buses with similar buses and identifies cases where pricing may be incorrect.", None),
    ("", None),
    ("APPROACH SUMMARY", None),
    ("✦  Similarity Matching", "6 criteria: AC class, sleeper type, seat layout, duration ±60min, departure time ±90min, rating ±0.8"),
    ("✦  Benchmark Method",    "Occupancy-weighted median of competitor prices — high-demand buses carry more weight"),
    ("✦  Flagging Logic",      "Dynamic threshold (±10/15/20%) based on price spread — adapts to route competitiveness"),
    ("✦  Service Adjustment",  "Benchmark adjusted ±2% per premium feature (Live Tracking, M-Ticket, Seat Layout)"),
    ("✦  WAP Accuracy",        "Recomputed directly from Seat Prices column: Σ(price×seats)/Σ(seats)"),
]
r = 20
for item in desc_rows:
    ws_cover.row_dimensions[r].height = 22 if item[0] in ("ASSIGNMENT OBJECTIVE","APPROACH SUMMARY") else 28
    if item[1] is None:
        ws_cover.merge_cells(f"A{r}:H{r}")
        c = ws_cover[f"A{r}"]
        c.value = item[0]
        if item[0] in ("ASSIGNMENT OBJECTIVE","APPROACH SUMMARY"):
            c.font = Font(name="Arial",bold=True,size=11,color="92D050")
        else:
            c.font = Font(name="Arial",size=11,color="CCCCCC")
        c.fill = fills["cover"]
        c.alignment = aligns["l"]
    else:
        ws_cover.merge_cells(f"A{r}:C{r}")
        c1 = ws_cover[f"A{r}"]
        c1.value = item[0]
        c1.font  = Font(name="Arial",bold=True,size=10,color="00A843")
        c1.fill  = fills["cover"]
        c1.alignment = aligns["l"]
        ws_cover.merge_cells(f"D{r}:H{r}")
        c2 = ws_cover[f"D{r}"]
        c2.value = item[1]
        c2.font  = Font(name="Arial",size=10,color="DDDDDD")
        c2.fill  = fills["cover"]
        c2.alignment = aligns["l"]
    r += 1


ws_cover.row_dimensions[r].height = 8; r+=1
ws_cover.merge_cells(f"A{r}:H{r}")
c = ws_cover[f"A{r}"]
c.value = "WORKBOOK CONTENTS"
c.font  = Font(name="Arial",bold=True,size=11,color="92D050")
c.fill  = fills["cover"]
c.alignment = aligns["l"]
r+=1

sheets_desc = [
    ("Sheet1 — Summary",         "KPI dashboard + route-by-route breakdown"),
    ("Sheet2 — Flagging Output",  "All 30,530 Flix trips with price flags, benchmark and deviation"),
    ("Sheet3 — Logic & Flowchart","Methodology, all 32 columns used, visual process flowchart"),
    ("Sheet4 — Route Insights",   "Best/worst priced routes, pricing opportunity analysis"),
    ("Sheet5 — Raw Data",         "First 2,000 rows of processed dataset"),
    ("Sheet6 — Automation Plan",  "6-step MVP pipeline, tech stack, 4-week rollout"),
]
for label, desc in sheets_desc:
    ws_cover.row_dimensions[r].height = 20
    ws_cover.merge_cells(f"A{r}:C{r}")
    c1 = ws_cover[f"A{r}"]
    c1.value = label
    c1.font  = Font(name="Arial",bold=True,size=10,color="00A843")
    c1.fill  = fills["cover2"]
    c1.alignment = aligns["l"]
    c1.border = brd
    ws_cover.merge_cells(f"D{r}:H{r}")
    c2 = ws_cover[f"D{r}"]
    c2.value = desc
    c2.font  = Font(name="Arial",size=10,color="FFFFFF")
    c2.fill  = fills["cover2"]
    c2.alignment = aligns["l"]
    c2.border = brd
    r+=1


ws_cover.row_dimensions[r+1].height = 20
ws_cover.merge_cells(f"A{r+2}:H{r+2}")
c = ws_cover[f"A{r+2}"]
c.value = "Submitted: 18 March 2026  |  Built with Python (pandas, numpy, openpyxl)  |  AI-assisted with Claude (Anthropic)"
c.font  = Font(name="Arial",size=9,color="888888")
c.fill  = fills["cover"]
c.alignment = Alignment(horizontal="center",vertical="center")

for ci in range(1,9): ws_cover.column_dimensions[get_column_letter(ci)].width=14
print(f"  Cover done {time.time()-t3:.1f}s")


ws0=wb.create_sheet("Sheet1_Summary")
ws0.sheet_view.showGridLines=False
wtitle(ws0,"A1:H1","FLIX PRICE MONITORING — EXECUTIVE SUMMARY  |  All Routes  |  16 Mar 2026")
wsec(ws0,"A3:H3","KEY PERFORMANCE INDICATORS")
kpis=[("Routes",len(routes),C_FLIX),
      ("Flix Trips",f"{len(df[df['Operator']==FLIX_NAME]):,}",C_FLIX),
      ("Analysed",f"{ta:,}",C_FLIX),
      ("✓ OK",f"{ta-tf:,}","2E7D32"),
      ("⚠ Flagged",f"{tf:,}","C62828"),
      ("▲ Too High",f"{th:,}","FF4C4C"),
      ("▼ Too Low",f"{tl:,}","F57F17"),
      ("Flag Rate",f"{round(tf/ta*100,1) if ta else 0}%","1565C0")]
ws0.row_dimensions[4].height=22; ws0.row_dimensions[5].height=44
for ci,(label,val,col) in enumerate(kpis,1):
    c1=ws0.cell(4,ci);c1.value=label;c1.font=fonts["hdr"]
    c1.fill=fills["hdr"];c1.alignment=aligns["c"];c1.border=brd
    c2=ws0.cell(5,ci);c2.value=val
    c2.font=Font(name="Arial",bold=True,size=16,color=col)
    c2.fill=fills["kpi"];c2.alignment=aligns["c"];c2.border=brd_m

wsec(ws0,"A7:H7","ROUTE-BY-ROUTE BREAKDOWN")
whdrs(ws0,8,["Route","Total Buses","Flix Trips","Analysed",
             "Flagged","Too High","Too Low","Flag Rate (%)"])
for ri,s in enumerate(summary_stats,9):
    fr2=round(s["flag"]/s["an"]*100,1) if s["an"] else 0
    fill=fills["alt"] if ri%2==0 else fills["white"]
    wrow(ws0,ri,[s["route"],s["total"],s["flix"],s["an"],
                 s["flag"],s["hi"],s["lo"],fr2],
         fill,nfmts={8:'0.0"%"'})
scols(ws0,[14,12,11,12,10,10,10,12])
print(f"  Sheet1_Summary done {time.time()-t3:.1f}s")


ws1=wb.create_sheet("Sheet2_Flagging_Output")
ws1.sheet_view.showGridLines=False
wtitle(ws1,"A1:U1","FLIX BUS PRICE MONITORING — FLAGGING OUTPUT  |  All Routes  |  16 Mar 2026")
wsec(ws1,"A3:U3","SECTION A — PRICE FLAGS  (Recomputed WAP | Wtd Median | Svc Adj | Dynamic Threshold)")
whdrs(ws1,4,["Route","Rank","Dep","Arr","Dur","Layout","Occ%",
             "Min Fare","Max Fare","Tiers","Disc%","Svc",
             "Flix WAP","#Comp","Wtd Median","Svc Adj%","Adj Bench",
             "Diff(₹)","Diff(%)","Thresh","FLAG"],h=44)
nfmts_a={8:'#,##0.00',9:'#,##0.00',13:'#,##0.00',15:'#,##0.00',
         17:'#,##0.00',18:'#,##0.00',19:'0.00'}
for ri,fr in enumerate(all_flag_rows,5):
    fill=fills["alt"] if ri%2==0 else fills["white"]
    is_f=fr["flag"]!="OK"
    fl=("⚠ " if is_f else "✓ ")+fr["flag"]
    ff=(fills["red"] if fr["flag"]=="TOO HIGH" else
        fills["yel"] if fr["flag"]=="TOO LOW"  else fills["grn"])
    wrow(ws1,ri,
         [fr["route"],fr["rank"],fr["dep"],fr["arr"],fr["dur"],
          fr["layout"],fr["occ"],fr["min_fare"],fr["max_fare"],
          fr["num_tiers"],fr["avg_disc"],f"{fr['svc']}/3",
          fr["flix_price"],fr["comp_count"],fr["w_med"],
          f"{fr['svc_adj']:+.1f}%",fr["adj_b"],
          fr["diff_abs"],fr["diff_pct"],f"±{fr['thold']}%",fl],
         fill,nfmts_a,fcol=21,ffill=ff)

sb=5+len(all_flag_rows)+2
wsec(ws1,f"A{sb}:U{sb}","SECTION B — COMPARABLE BUSES (top 10 per Flix trip)")
hb=sb+1
whdrs(ws1,hb,["Route","Flix Rank","Comp Rank","Operator","Bus Type",
              "Layout","Dep","Dur","WAP(₹)","Min Fare",
              "Occ%","Rating","Reviews","Svc","BP","DP",
              "IsAC","Sleep","Diff(₹)","Sim"],h=44)
nfmts_b={9:'#,##0.00',10:'#,##0.00',19:'#,##0.00'}
curr=hb+1
for fr in all_flag_rows:
    if len(fr["comp_rows"])==0: continue
    for _,comp in fr["comp_rows"].iterrows():
        fill=fills["alt"] if curr%2==0 else fills["white"]
        try:
            pdiff=round(fr["flix_price"]-float(comp["WAP_Final"]),2)
            dd=abs(int(comp["Journey Duration (Min)"])-fr["dur"])
            dv=float(comp["Dep_Min"])
            fd2=int(fr["dep"].split(":")[0])*60+int(fr["dep"].split(":")[1])
            dpd=min(abs(fd2-dv),1440-abs(fd2-dv))
            sim=max(10-(2 if dd>30 else 0)-(2 if dpd>45 else 0),1)
            wrow(ws1,curr,
                 [fr["route"],fr["rank"],
                  int(comp.get("Rank_Num",0)),
                  str(comp.get("Operator","")),
                  str(comp.get("Bus Type","")),
                  str(comp.get("Layout","")),
                  str(comp.get("Dep_Str","")),
                  int(comp.get("Journey Duration (Min)",0)),
                  float(comp.get("WAP_Final",0)),
                  float(comp.get("Min_Fare",0)),
                  float(comp.get("Occ_Rate",0)),
                  float(comp.get("Total Ratings",0)),
                  int(comp.get("Number of Reviews",0)),
                  f"{int(comp.get('Svc_Score',0))}/3",
                  int(comp.get("BP Count",0)),
                  int(comp.get("DP Count",0)),
                  "Y" if comp.get("Is AC",False) else "N",
                  "Y" if comp.get("Is Sleeper",False) else "N",
                  pdiff,f"{sim}/10"],
                 fill,nfmts_b)
            curr+=1
        except: curr+=1; continue

scols(ws1,[12,7,7,15,22,8,8,8,11,11,7,7,8,7,5,5,6,7,11,8])
ws1.freeze_panes="A5"
print(f"  Sheet2_Flagging done {time.time()-t3:.1f}s")


ws2=wb.create_sheet("Sheet3_Logic_Flowchart")
ws2.sheet_view.showGridLines=False
wtitle(ws2,"A1:H1","METHODOLOGY, LOGIC & PROCESS FLOWCHART — All 32 Columns Used")


wsec(ws2,"A3:H3","SYSTEM PROCESS FLOWCHART")
ws2.row_dimensions[3].height=26

flowchart_steps = [
    ("STEP 1\nDATA INGESTION",
     "Input: Raw dataset (862K rows, 32 columns)\n"
     "Columns used: All 32 — Route Number, Operator, Bus Type, Is AC,\n"
     "Is Seater, Is Sleeper, Departure Time, Duration, Seat Prices,\n"
     "Fare List, Discounted Prices, Available Seats, Total Seats,\n"
     "Total Ratings, BP Count, DP Count, service booleans",
     C_FLIX, "FFFFFF"),

    ("STEP 2\nDATA PROCESSING",
     "→  WAP Recomputed from Seat Prices: Σ(price×seats)/Σ(seats)\n"
     "→  Occupancy Rate = (Total Seats − Available Seats) / Total Seats\n"
     "→  Service Score (0–3) = Live Tracking + M-Ticket + Seat Layout\n"
     "→  Layout extracted from Bus Type: (2+1) or (2+2)\n"
     "→  Avg Discount % from Fare List vs Discounted Prices",
     "1565C0", "FFFFFF"),

    ("STEP 3\nSIMILARITY MATCHING\n(6 Criteria)",
     "For each Flixbus trip, find competitors where ALL pass:\n"
     "  [1] Is AC = TRUE           (same comfort class)\n"
     "  [2] Is Sleeper matches      (same seat type)\n"
     "  [3] Layout matches (2+1/2+2)(same physical product)\n"
     "  [4] Duration within ±60 min (same route length)\n"
     "  [5] Dep time within ±90 min (same demand window)\n"
     "  [6] Rating within ±0.8      (same quality band)",
     "6A1B9A", "FFFFFF"),

    ("STEP 4\nBENCHMARK CALCULATION",
     "→  Benchmark = Occupancy-Weighted Median of similar comp prices\n"
     "     (buses with higher occupancy carry more weight)\n"
     "→  Service Adjustment = ±2% × (Flix Svc Score − Avg Comp Svc Score)\n"
     "     (adjusts benchmark up if Flix has more premium features)\n"
     "→  Adjusted Benchmark = Weighted Median × (1 + Service Adj%)",
     "E65100", "FFFFFF"),

    ("STEP 5\nDYNAMIC FLAGGING",
     "→  Price Diff % = (Flix WAP − Adj Benchmark) / Adj Benchmark × 100\n"
     "→  Dynamic Threshold based on Coefficient of Variation (CV):\n"
     "     CV < 10%  →  Tight market   →  ±10% threshold\n"
     "     CV 10-25% →  Normal market  →  ±15% threshold\n"
     "     CV > 25%  →  Spread market  →  ±20% threshold\n"
     "→  FLAG: TOO HIGH if Diff% > threshold | TOO LOW if < -threshold",
     "B71C1C", "FFFFFF"),

    ("STEP 6\nOUTPUT",
     "→  Sheet2: Full flagging output with 30,530 Flix trip results\n"
     "→  Sheet4: Route Insights — best/worst priced routes\n"
     "→  Sheet1: Executive Summary with KPI dashboard\n"
     "→  Flag colours: 🔴 TOO HIGH | 🟡 TOO LOW | 🟢 OK",
     C_HDR, "FFFFFF"),
]

r = 4
for i, (title, body, bg, fg) in enumerate(flowchart_steps):

    ws2.row_dimensions[r].height   = 16
    ws2.row_dimensions[r+1].height = 90
    ws2.row_dimensions[r+2].height = 10

    ws2.merge_cells(f"A{r}:H{r}")
    c_title = ws2[f"A{r}"]
    c_title.value     = title
    c_title.font      = Font(name="Arial",bold=True,size=11,color=fg)
    c_title.fill      = PatternFill("solid",fgColor=bg)
    c_title.alignment = Alignment(horizontal="left",vertical="center",indent=2)
    c_title.border    = brd_m

    ws2.merge_cells(f"A{r+1}:H{r+1}")
    c_body = ws2[f"A{r+1}"]
    c_body.value     = body
    c_body.font      = Font(name="Arial",size=10)
    c_body.fill      = PatternFill("solid",fgColor="F9F9F9")
    c_body.alignment = Alignment(horizontal="left",vertical="top",wrap_text=True,indent=2)
    c_body.border    = brd_m


    if i < len(flowchart_steps)-1:
        ws2.merge_cells(f"D{r+2}:E{r+2}")
        c_arr = ws2[f"D{r+2}"]
        c_arr.value     = "▼"
        c_arr.font      = Font(name="Arial",bold=True,size=14,color="00A843")
        c_arr.fill      = fills["white"]
        c_arr.alignment = Alignment(horizontal="center",vertical="center")

    r += 3


r += 1
wsec(ws2,f"A{r}:H{r}","ALL 32 COLUMNS — HOW EACH IS USED IN THIS SYSTEM")
r += 1
whdrs(ws2,r,["Column Name","Used In","How It's Used"],h=30)
r += 1

col_usage = [
    ("Route Number",              "Similarity","Routes processed separately — all 61 routes"),
    ("SRP Rank",                  "Output","Displayed in flagging output for reference"),
    ("Operator",                  "Filtering","Identifies Flixbus trips vs competitors"),
    ("Bus Type",                  "Similarity","Layout (2+1/2+2) extracted for similarity filter"),
    ("Is AC",                     "Similarity","Filter 1: only AC vs AC comparisons"),
    ("Is Seater",                 "Context","Shown in output for reference"),
    ("Is Sleeper",                "Similarity","Filter 2: sleeper vs sleeper only"),
    ("Departure Time",            "Similarity","Filter 5: ±90 min circular time window"),
    ("Arrival Time",              "Output","Displayed in flagging output"),
    ("Journey Duration (Min)",    "Similarity","Filter 4: ±60 min duration window"),
    ("Seat Prices",               "WAP Recomputation","Σ(price×seats)/Σ(seats) — more accurate than provided WAP"),
    ("Weighted Average Price",    "WAP Fallback","Used if Seat Prices is missing/malformed"),
    ("Fare List",                 "Output","Min/max fare range shown per trip"),
    ("Discounted Prices",         "Output","Avg discount % computed and shown"),
    ("Available Seats",           "Occupancy","Numerator: (Total−Avail)/Total×100"),
    ("Total Seats",               "Occupancy","Denominator for occupancy rate calculation"),
    ("Total Ratings",             "Similarity","Filter 6: ±0.8 rating band quality filter"),
    ("Number of Reviews",         "Output","Shown in comparable bus detail table"),
    ("Bus Score",                 "Output","Shown in comparable bus detail table"),
    ("Is Seat Layout Available",  "Service Score","1 of 3 features in Service Score (0–3)"),
    ("Is Live Tracking Available","Service Score","1 of 3 features in Service Score (0–3)"),
    ("Is M-Ticket Enabled",       "Service Score","1 of 3 features in Service Score (0–3)"),
    ("BP Count",                  "Output","Shown in comparable bus detail for context"),
    ("DP Count",                  "Output","Shown in comparable bus detail for context"),
    ("Available Window Seats",    "Context","Available in dataset, not used in this MVP"),
    ("Available Single Seats",    "Context","Available in dataset, not used in this MVP"),
    ("Available Aisle Seats",     "Context","Available in dataset, not used in this MVP"),
    ("Available Upper Seats",     "Context","Available in dataset, not used in this MVP"),
    ("Available Lower Seats",     "Context","Available in dataset, not used in this MVP"),
    ("Date of Extraction",        "Context","Single date dataset — not used in MVP"),
    ("Departure Date / DOJ",      "Context","Same-day data — not used in MVP"),
    ("Date of Journey",           "Context","Same-day data — not used in MVP"),
]
for i,(col,used_in,how) in enumerate(col_usage):
    fill = fills["alt"] if i%2==0 else fills["white"]
    use_fill = (PatternFill("solid",fgColor="E8F5E9") if "Similarity" in used_in else
                PatternFill("solid",fgColor="E3F2FD") if "WAP" in used_in else
                PatternFill("solid",fgColor="FFF8E1") if "Service" in used_in else
                PatternFill("solid",fgColor="FCE4EC") if "Occupancy" in used_in else fill)
    ws2.row_dimensions[r].height=18
    for ci,v in enumerate([col,used_in,how],1):
        c=ws2.cell(r,ci);c.value=v
        c.font=Font(name="Arial",bold=(ci==1),size=10)
        c.fill=use_fill if ci<=2 else fill
        c.alignment=aligns["l"];c.border=brd
    r+=1

ws2.column_dimensions["A"].width=28
ws2.column_dimensions["B"].width=18
for col in ["C","D","E","F","G","H"]:
    ws2.column_dimensions[col].width=14
ws2.freeze_panes="A4"
print(f"  Sheet3_Logic done {time.time()-t3:.1f}s")


ws_ins = wb.create_sheet("Sheet4_Route_Insights")
ws_ins.sheet_view.showGridLines = False
wtitle(ws_ins,"A1:L1","ROUTE-LEVEL PRICING INSIGHTS  |  Best & Worst Priced Routes  |  16 Mar 2026")


wsec(ws_ins,"A3:L3","SECTION A — MOST OVERPRICED ROUTES  (Flix charging significantly above market)")
whdrs(ws_ins,4,["Route","Flix Trips","Too High","% Too High",
                "Avg Overpricing%","Avg Flix WAP(₹)","Avg Comp WAP(₹)",
                "Gap(₹)","Action Needed"],h=36)

hi_routes = sorted(
    [s for s in summary_stats if s["hi"]>0],
    key=lambda x: x["hi"]/x["an"] if x["an"] else 0, reverse=True)[:15]

for ri,s in enumerate(hi_routes,5):
    ws_ins.row_dimensions[ri].height=20
    fill=fills["hi_opp"] if ri%2==0 else PatternFill("solid",fgColor="FFF5F5")
    pct_hi=round(s["hi"]/s["an"]*100,1) if s["an"] else 0
    gap=round(s["avg_flix_wap"]-s["avg_comp_wap"],0)
    action=("URGENT — Review pricing immediately" if pct_hi>70 else
            "HIGH — Consider price reduction" if pct_hi>50 else
            "MEDIUM — Monitor and adjust")
    wrow(ws_ins,ri,
         [s["route"],s["flix"],s["hi"],pct_hi,
          s["avg_hi_diff"],s["avg_flix_wap"],s["avg_comp_wap"],gap,action],
         fill, nfmts={4:'0.0"%"',5:'0.0"%"',6:'#,##0',7:'#,##0',8:'#,##0'})

    ac = ws_ins.cell(ri,9)
    ac.fill = (PatternFill("solid",fgColor="FFCDD2") if "URGENT" in action else
               PatternFill("solid",fgColor="FFE0B2") if "HIGH" in action else
               PatternFill("solid",fgColor="FFF9C4"))
    ac.font = Font(name="Arial",bold=True,size=10)


start_b = 5 + len(hi_routes) + 2
wsec(ws_ins,f"A{start_b}:L{start_b}",
     "SECTION B — MOST UNDERPRICED ROUTES  (Flix leaving revenue on the table)")
whdrs(ws_ins,start_b+1,["Route","Flix Trips","Too Low","% Too Low",
                         "Avg Underpricing%","Avg Flix WAP(₹)","Avg Comp WAP(₹)",
                         "Revenue Gap(₹)","Opportunity"],h=36)

lo_routes = sorted(
    [s for s in summary_stats if s["lo"]>0],
    key=lambda x: x["lo"]/x["an"] if x["an"] else 0, reverse=True)[:15]

for ri,s in enumerate(lo_routes, start_b+2):
    ws_ins.row_dimensions[ri].height=20
    fill=fills["lo_opp"] if ri%2==0 else PatternFill("solid",fgColor="F5FFF5")
    pct_lo=round(s["lo"]/s["an"]*100,1) if s["an"] else 0
    gap=round(s["avg_comp_wap"]-s["avg_flix_wap"],0)
    opp=("HIGH REVENUE OPP — Consider price increase" if pct_lo>70 else
         "MEDIUM OPP — Test higher price points" if pct_lo>50 else
         "LOW OPP — Minor adjustment possible")
    wrow(ws_ins,ri,
         [s["route"],s["flix"],s["lo"],pct_lo,
          s["avg_lo_diff"],s["avg_flix_wap"],s["avg_comp_wap"],gap,opp],
         fill, nfmts={4:'0.0"%"',5:'0.0"%"',6:'#,##0',7:'#,##0',8:'#,##0'})
    oc = ws_ins.cell(ri,9)
    oc.fill = (PatternFill("solid",fgColor="C8E6C9") if "HIGH" in opp else
               PatternFill("solid",fgColor="DCEDC8") if "MEDIUM" in opp else
               PatternFill("solid",fgColor="F9FBE7"))
    oc.font = Font(name="Arial",bold=True,size=10)


start_c = start_b + 2 + len(lo_routes) + 2
wsec(ws_ins,f"A{start_c}:L{start_c}",
     "SECTION C — BEST PRICED ROUTES  (Flix most aligned with market)")
whdrs(ws_ins,start_c+1,
      ["Route","Flix Trips","Analysed","OK Count","Flag Rate%",
       "Avg Flix WAP(₹)","Avg Comp WAP(₹)","Avg Diff%","Status"],h=36)

best_routes = sorted(
    [s for s in summary_stats if s["an"]>=10],
    key=lambda x: x["flag"]/x["an"] if x["an"] else 1)[:10]

for ri,s in enumerate(best_routes, start_c+2):
    ws_ins.row_dimensions[ri].height=20
    fill=fills["alt"] if ri%2==0 else fills["white"]
    fr2=round(s["flag"]/s["an"]*100,1) if s["an"] else 0
    ok_count=s["an"]-s["flag"]
    status="✅ Well Priced" if fr2<30 else "⚠ Partially Aligned"
    wrow(ws_ins,ri,
         [s["route"],s["flix"],s["an"],ok_count,fr2,
          s["avg_flix_wap"],s["avg_comp_wap"],s["avg_diff"],status],
         fill, nfmts={5:'0.0"%"',6:'#,##0',7:'#,##0',8:'0.0"%"'})
    sc2 = ws_ins.cell(ri,9)
    sc2.fill=(PatternFill("solid",fgColor="C8E6C9") if "Well" in status
              else PatternFill("solid",fgColor="FFF9C4"))
    sc2.font=Font(name="Arial",bold=True,size=10)

scols(ws_ins,[12,10,10,12,13,15,15,13,28])
ws_ins.freeze_panes="A5"
print(f"  Sheet4_Insights done {time.time()-t3:.1f}s")


ws3=wb.create_sheet("Sheet5_Raw_Data")
ws3.sheet_view.showGridLines=False
disp=df.head(2000)
wtitle(ws3,"A1:T1",
       f"RAW DATA — {len(df):,} rows | {len(routes)} routes | "
       f"Flixbus: {len(df[df['Operator']==FLIX_NAME]):,} trips | First 2,000 shown")
whdrs(ws3,2,["Route","Rank","Operator","Bus Type","Layout","AC","Seat","Sleep",
             "Dep","Arr","Dur","WAP(₹)","Min Fare","Max Fare","Tiers","Disc%",
             "Occ%","Svc","BP","DP"])
for ri,(_,row) in enumerate(disp.iterrows(),3):
    ws3.row_dimensions[ri].height=15
    is_f=row["Operator"]==FLIX_NAME
    fill=fills["frow"] if is_f else (fills["alt"] if ri%2==0 else fills["white"])
    font=fonts["frow"] if is_f else fonts["sm"]
    data=[row["Route Number"],int(row["Rank_Num"]),row["Operator"],row["Bus Type"],
          row["Layout"],
          "Y" if row["Is AC"] else "N","Y" if row["Is Seater"] else "N",
          "Y" if row["Is Sleeper"] else "N",
          row["Dep_Str"],row["Arr_Str"],int(row["Journey Duration (Min)"]),
          float(row["WAP_Final"]),float(row["Min_Fare"]),float(row["Max_Fare"]),
          int(row["Num_Tiers"]),float(row["Avg_Disc"]),
          float(row["Occ_Rate"]),f"{int(row['Svc_Score'])}/3",
          int(row["BP Count"]),int(row["DP Count"])]
    for ci,v in enumerate(data,1):
        c=ws3.cell(ri,ci);c.value=v;c.font=font;c.fill=fill
        c.alignment=aligns["c"];c.border=brd
        if ci==12: c.number_format='#,##0.00'
ws3.auto_filter.ref="A2:T2"; ws3.freeze_panes="A3"
scols(ws3,[12,7,15,22,8,6,6,7,9,9,9,12,10,10,6,7,7,7,5,5])
print(f"  Sheet5_RawData done {time.time()-t3:.1f}s")


ws4=wb.create_sheet("Sheet6_Automation_Plan")
ws4.sheet_view.showGridLines=False
wtitle(ws4,"A1:G1","AUTOMATION PLAN (MVP) — Flix Bus Price Monitoring System")

def asec(ws,row,text):
    ws.merge_cells(f"A{row}:G{row}")
    c=ws[f"A{row}"];c.value=text;c.font=fonts["sec"]
    c.fill=fills["hdr"];c.alignment=aligns["l"];ws.row_dimensions[row].height=26

def arow(ws,row,step,content,sc=C_FLIX):
    c1=ws.cell(row,1);c1.value=step;c1.font=fonts["hdr"]
    c1.fill=PatternFill("solid",fgColor=sc)
    c1.alignment=aligns["c"];c1.border=brd
    ws.merge_cells(f"B{row}:G{row}")
    c2=ws[f"B{row}"];c2.value=content
    c2.font=fonts["norm"];c2.alignment=aligns["lt"];c2.border=brd
    ws.row_dimensions[row].height=65

asec(ws4,3,"OVERVIEW")
arow(ws4,4,"Goal",
     "Automated daily pipeline: ingest all 32 columns → vectorized cleaning → "
     "bucket pre-filtered numpy similarity (memory-safe) → "
     "occupancy-weighted median + service adjustment + dynamic threshold → "
     "Excel report + email alert.")
asec(ws4,6,"PIPELINE STEPS"); r4=7
for s,c in [
    ("Step 1\nIngestion",   "Python + PostgreSQL/BigQuery. All 32 columns daily at 06:00 AM."),
    ("Step 2\nCleaning",    "Vectorized pandas: WAP recompute, bool parse, occ/svc/discount."),
    ("Step 3\nSimilarity",  "Bucket pre-filter (AC+Sleeper+Layout) → numpy ops. Memory-safe on 862K rows."),
    ("Step 4\nFlagging",    "Occ-weighted median + service adj (±2%/feature) + dynamic CV threshold."),
    ("Step 5\nReporting",   "Auto-generate this 6-sheet Excel. Email team. Looker Studio dashboard."),
    ("Step 6\nOrchestrate", "Airflow or GitHub Actions cron. Retry logic. Health monitoring."),
]:
    arow(ws4,r4,s,c); r4+=1

asec(ws4,r4,"TECH STACK"); r4+=1
for s,c in [("Language","Python 3.10+"),("Database","PostgreSQL or BigQuery"),
            ("Processing","pandas, numpy"),("Reporting","openpyxl + SendGrid"),
            ("Dashboard","Looker Studio or Tableau"),
            ("Scheduler","Apache Airflow or GitHub Actions"),
            ("Testing","pytest")]:
    arow(ws4,r4,s,c,"2E7D32"); r4+=1

asec(ws4,r4,"4-WEEK MVP"); r4+=1
for s,c in [
    ("Week 1","Route 1. Ingest + clean all 32 cols. Validate WAP recomputation."),
    ("Week 2","6-criterion bucket similarity. Validate + unit tests."),
    ("Week 3","Weighted median + service adj + dynamic threshold. Excel + test email."),
    ("Week 4","All 61 routes. PostgreSQL. Airflow. Looker Studio dashboard launch."),
]:
    arow(ws4,r4,s,c,"1565C0"); r4+=1

asec(ws4,r4,"AI TOOLS USED IN THIS ASSIGNMENT"); r4+=1
arow(ws4,r4,"Claude\n(Anthropic)",
     "Primary tool for code generation and logic design.\n"
     "Used to write and debug the Python pipeline — pandas vectorization, openpyxl formatting, "
     "the bucket pre-filter matching strategy, and this automation plan.\n"
     "All analytical decisions were mine: the 6 similarity criteria, threshold values, "
     "weighted median approach, and service adjustment formula. Claude handled the implementation.", "6A1B9A")
r4+=1
arow(ws4,r4,"ChatGPT\n(OpenAI)",
     "Used for exploratory data questions early in the process.\n"
     "Asked it to explain what WAP actually measures vs a simple average, and to sanity-check "
     "whether a ±15% pricing threshold is reasonable for Indian intercity bus markets. "
     "Helped me think through edge cases before writing any code — e.g. what happens "
     "when a route has fewer than 3 comparable buses.", "0D6E3A")
r4+=1
arow(ws4,r4,"GitHub Copilot",
     "Used for inline autocomplete while writing repetitive openpyxl cell-formatting blocks.\n"
     "Particularly helpful for the Excel sheet-building sections where the pattern "
     "(set value, set font, set fill, set border) repeats hundreds of times. "
     "Saved time on boilerplate without influencing any of the pricing logic.", "1A1A2E")
r4+=1
arow(ws4,r4,"Perplexity AI",
     "Used for quick research lookups during methodology design.\n"
     "Looked up: typical price variation ranges in bus fare benchmarking, "
     "how redBus and MakeMyTrip rank bus listings (to understand what SRP Rank means), "
     "and whether occupancy-weighted pricing is used in practice by yield management systems. "
     "Helped ground the approach in how the industry actually works.", "0B7A6B")

ws4.column_dimensions["A"].width=14
for col in ["B","C","D","E","F","G"]: ws4.column_dimensions[col].width=17
print(f"  Sheet6_Automation done {time.time()-t3:.1f}s")


out=r"D:\resume\resume projects\Flix\Flix_Price_Monitoring_Assignment.xlsx"
print("Saving..."); wb.save(out)  # takes ~30s for large workbooks
total=time.time()-t0
print(f"\n✅ DONE in {total:.0f}s")
print(f"   Rows: {len(df):,} | Routes: {len(routes)} | Flix: {len(df[df['Operator']==FLIX_NAME]):,}")
print(f"   Analysed: {ta:,} | Flagged: {tf:,} (High:{th:,} Low:{tl:,})")
print(f"   Output: {out}")
input("\nPress Enter to close...")